import io
import uuid
from collections.abc import Sequence

import openpyxl
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.exceptions import UnprocessableError
from app.models.enums import InstallmentStatus
from app.models.policy import Policy
from app.models.premium_installment import PremiumInstallment
from app.schemas.payment_batch import (
    BulkPaymentFileResolution,
    BulkPaymentResolvedRow,
    BulkPaymentUnmatchedRow,
)

_HEADER_ALIASES = {
    "reference number",
    "reference",
    "debit note",
    "debit note number",
    "ref no",
    "ref number",
}


def _find_reference_column(header_row: Sequence[object]) -> int | None:
    for idx, cell in enumerate(header_row):
        if isinstance(cell, str) and cell.strip().lower() in _HEADER_ALIASES:
            return idx
    return None


def build_reference_template_workbook() -> bytes:
    """A minimal, downloadable .xlsx a broker can fill in and re-upload
    through the same panel -- one column, the canonical header this
    module's own _HEADER_ALIASES already accepts, plus example rows so the
    expected format is unambiguous without needing separate instructions.
    """
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert sheet is not None  # a freshly-constructed Workbook always has one active sheet
    sheet.title = "Reference numbers"
    sheet.append(["Reference Number"])
    sheet.append(["DN-2026-0001"])
    sheet.append(["DN-2026-0002"])

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


async def resolve_installments_from_excel(
    db: AsyncSession,
    *,
    file_bytes: bytes,
    broker_id: uuid.UUID,
    max_rows: int,
) -> BulkPaymentFileResolution:
    """Pure lookup by PremiumInstallment.reference_number -- never assigns a
    reference number as a side effect (that's PATCH
    /installments/{id}/reference-number, a broker/insurer sets it
    explicitly before ever uploading a file). Never touches
    Payment/PaymentBatch -- the caller submits the resolved installment_ids
    through the existing, unmodified POST /payments/bulk once the broker
    confirms the preview.

    Ownership is checked by broker_id alone, not insurance_company_id -- a
    broker may work with several insurers at once, and this is a preview/
    match step the broker reviews and edits before submission, not the point
    where the single-insurer-per-batch rule needs to bite (that's enforced
    in bulk_payment_service._load_payable_installments, at actual payment
    creation time).
    """
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise UnprocessableError("file is not a valid .xlsx workbook") from exc

    sheet = workbook.worksheets[0]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return BulkPaymentFileResolution(
            total_rows=0, matched=[], unmatched=[], file_error="file is empty"
        )

    ref_col = _find_reference_column(rows[0])
    if ref_col is None:
        return BulkPaymentFileResolution(
            total_rows=0,
            matched=[],
            unmatched=[],
            file_error="no recognizable reference-number column found in the header row",
        )

    data_rows = rows[1:]
    if len(data_rows) > max_rows:
        raise UnprocessableError(f"file has {len(data_rows)} rows, exceeding the cap of {max_rows}")

    # Batch-resolve rather than one query per row (300+ row target) -- same
    # two-query shape bulk_payment_service._load_payable_installments uses,
    # which an earlier load test hardened against the equivalent N+1.
    row_references: list[tuple[int, str]] = []
    unmatched: list[BulkPaymentUnmatchedRow] = []
    for row_number, row in enumerate(data_rows, start=2):  # spreadsheet row 2 = first data row
        raw_ref = row[ref_col] if ref_col < len(row) else None
        reference_number = str(raw_ref).strip() if raw_ref not in (None, "") else None
        if reference_number is None:
            unmatched.append(
                BulkPaymentUnmatchedRow(
                    row_number=row_number, reference_number=None, reason="missing reference number"
                )
            )
            continue
        row_references.append((row_number, reference_number))

    lookup_values = {ref.lower() for _, ref in row_references}
    installments_by_ref: dict[str, list[PremiumInstallment]] = {}
    if lookup_values:
        candidate_installments = (
            await db.scalars(
                select(PremiumInstallment).where(
                    func.lower(PremiumInstallment.reference_number).in_(lookup_values)
                )
            )
        ).all()
        for installment in candidate_installments:
            if installment.reference_number is None:
                continue
            key = installment.reference_number.lower()
            installments_by_ref.setdefault(key, []).append(installment)

    policy_ids = {
        installment.policy_id
        for installments in installments_by_ref.values()
        for installment in installments
    }
    policies_by_id: dict[uuid.UUID, Policy] = {}
    if policy_ids:
        policy_rows = (await db.scalars(select(Policy).where(Policy.id.in_(policy_ids)))).all()
        policies_by_id = {p.id: p for p in policy_rows}

    matched: list[BulkPaymentResolvedRow] = []
    seen_installment_ids: set[uuid.UUID] = set()
    for row_number, reference_number in row_references:
        candidates = installments_by_ref.get(reference_number.lower(), [])
        if len(candidates) == 0:
            unmatched.append(
                BulkPaymentUnmatchedRow(
                    row_number=row_number,
                    reference_number=reference_number,
                    reason="no installment found with this reference number",
                )
            )
            continue
        if len(candidates) > 1:
            unmatched.append(
                BulkPaymentUnmatchedRow(
                    row_number=row_number,
                    reference_number=reference_number,
                    reason=(
                        f"reference number matches {len(candidates)} installments -- "
                        "ambiguous, contact support"
                    ),
                )
            )
            continue

        installment = candidates[0]
        if installment.id in seen_installment_ids:
            unmatched.append(
                BulkPaymentUnmatchedRow(
                    row_number=row_number,
                    reference_number=reference_number,
                    reason="duplicate row -- already matched earlier in this file",
                )
            )
            continue

        policy = policies_by_id.get(installment.policy_id)
        if policy is None or policy.broker_id != broker_id:
            unmatched.append(
                BulkPaymentUnmatchedRow(
                    row_number=row_number,
                    reference_number=reference_number,
                    reason="no installment found with this reference number",
                )
            )
            continue
        if installment.status not in (InstallmentStatus.DUE.value, InstallmentStatus.OVERDUE.value):
            unmatched.append(
                BulkPaymentUnmatchedRow(
                    row_number=row_number,
                    reference_number=reference_number,
                    reason=f"installment already {installment.status}",
                )
            )
            continue

        seen_installment_ids.add(installment.id)
        matched.append(
            BulkPaymentResolvedRow(
                row_number=row_number,
                installment_id=installment.id,
                reference_number=reference_number,
                amount_kobo=installment.amount_kobo,
                policy_id=installment.policy_id,
            )
        )

    return BulkPaymentFileResolution(
        total_rows=len(data_rows), matched=matched, unmatched=unmatched, file_error=None
    )
